import openpyxl

class XLSXTable:
    def __init__(self, tablePath, currentSheet='Sheet1', lastRecord='A1', lastIndex=1, lastColumn='A'):
        #TODO
        self.tablePath = tablePath
        self.currentSheet = currentSheet
        self.lastRecord = lastRecord
        self.lastIndex = lastIndex
        self.lastColumn = lastColumn
        self.sheets = 0

        self.workBook = None
        self.workSheet = None
    
    def createTable(self):
        #TODO
        self.workBook = openpyxl.Workbook()
        self.saveTable()

    def loadTable(self):
        #TODO
        self.workBook = openpyxl.load_workbook(self.tablePath)
    
    def saveTable(self):
        self.workBook.save(self.tablePath)

    def closeTable(self):
        self.workBook.close()

    def createNewSheet(self, sheetName: str):
        self.workBook.create_sheet(sheetName)
        self.saveTable()

    def setColumnWidth(self, column, width):
        self.workSheet.column_dimensions[column].width = width

    def setValueOnSheet(self, record: str, value: any):
        self.workSheet[record].value = value
        self.lastRecord = record
    
    def getValueOnSheet(self, record):
        return self.workSheet[record].value

    def addSheet(self):
        self.sheets += 1
    

    #set and get on self.*
    def setSheets(self, sheets: int):
        self.sheets = sheets

    def getSheets(self):
        return self.sheets

    def setLastRecord(self, lastRecord: str):
        self.lastRecord = lastRecord

    def getLastRecord(self):
        return self.lastRecord

    def setLastIndex(self, lastIndex: int):
        self.lastIndex = lastIndex

    def getLastIndex(self):
        return self.lastIndex
    
    def setLastColumn(self, lastColumn: str):
        self.lastColumn = lastColumn

    def getLastColumn(self):
        return self.lastColumn

    def setCurrentSheet(self, currentSheet: str):
        self.currentSheet = currentSheet
        self.workSheet = self.workBook.get_sheet_by_name(self.currentSheet)

    def getCurrentSheet(self):
        return self.currentSheet

    def setTablePath(self, tablePath: str):
        self.tablePath = tablePath

    def getTablePath(self):
        return self.tablePath

